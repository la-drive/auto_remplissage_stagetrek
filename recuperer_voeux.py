"""
recuperer_voeux.py
------------------
Ouvre le classement de stages de l'étudiant sur stagetrek.univ-tours.fr,
récupère le classement depuis la page /mon-stage/{id}#preferences,
et remplit le modèle Excel modele_voeux_campagne.xlsx avec les rangs.
"""

import sys
import os
import re
import traceback
import shutil
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

# ── Navigateur Playwright ─────────────────────────────────────────────────────
_DOSSIER_NAVIGATEURS = str(
    Path.home() / ".cache" / "auto-voeux-stagetrek" / "playwright-browsers"
)
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = _DOSSIER_NAVIGATEURS

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from bs4 import BeautifulSoup
import openpyxl


# ── Utilitaires console ───────────────────────────────────────────────────────

def pause(message="\nAppuie sur Entrée pour continuer..."):
    try:
        input(message)
    except EOFError:
        pass


def pause_avant_fermeture(message="\nAppuie sur Entrée pour fermer cette fenêtre..."):
    try:
        input(message)
    except EOFError:
        pass


# ── Installation du navigateur ────────────────────────────────────────────────

def ensure_browser_installed():
    print("Vérification du navigateur nécessaire (Chromium)...")
    try:
        from playwright.__main__ import main as playwright_cli_main
    except Exception as e:
        raise RuntimeError(f"Impossible de préparer Playwright : {e}")

    old_argv = sys.argv
    exit_code = 0
    try:
        sys.argv = ["playwright", "install", "chromium"]
        try:
            playwright_cli_main()
        except SystemExit as e:
            exit_code = e.code
    finally:
        sys.argv = old_argv

    if exit_code not in (0, None):
        raise RuntimeError("L'installation du navigateur a échoué.")


# ── Sélection du modèle Excel via boîte de dialogue ──────────────────────────

def choisir_modele_excel() -> Path:
    """
    Ouvre une boîte de dialogue pour que l'étudiant sélectionne
    le fichier modele_voeux_campagne.xlsx.
    Retourne le chemin sélectionné, ou lève SystemExit si annulé.
    """
    # Fenêtre Tk cachée (juste pour la boîte de dialogue)
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    print("Une fenêtre de sélection de fichier va s'ouvrir...")
    print("Sélectionne le fichier 'modele_voeux_campagne.xlsx'.\n")

    chemin = filedialog.askopenfilename(
        title="Sélectionne le fichier modèle de vœux",
        filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
    )

    root.destroy()

    if not chemin:
        print("❌ Aucun fichier sélectionné. Le programme va se fermer.")
        pause_avant_fermeture()
        sys.exit(0)

    return Path(chemin)


# ── Récupération du classement depuis la page HTML ────────────────────────────

def extraire_classement(html: str) -> dict[int, str]:
    """
    Analyse le HTML de la page /mon-stage/{id} et extrait le classement
    depuis l'onglet 'Préférences' (tableau #liste-preferences-*).

    Retourne un dict {rang: intitulé_terrain_principal}.

    Structure HTML cible :
        <table id="liste-preferences-...">
          <tbody>
            <tr>
              <td class="hidden">1</td>        <!-- rang (tri interne) -->
              <td class="text-center">1</td>   <!-- rang affiché       -->
              <td>CODE - INTITULE DU TERRAIN</td>
              <td>...</td>                     <!-- terrain secondaire  -->
              <td>...</td>                     <!-- demande             -->
            </tr>
            ...
          </tbody>
        </table>
    """
    soup = BeautifulSoup(html, "html.parser")

    # Chercher le tableau dont l'id commence par "liste-preferences-"
    tableau = soup.find("table", id=re.compile(r"^liste-preferences-"))
    if not tableau:
        return {}

    classement = {}
    for tr in tableau.find("tbody").find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue

        # td[0] = hidden (rang entier pour tri), td[1] = rang affiché
        try:
            rang = int(tds[1].get_text(strip=True))
        except ValueError:
            continue

        # td[2] = terrain principal : "CODE - INTITULE"
        terrain_brut = tds[2].get_text(separator=" ", strip=True)

        # Supprimer le code de discipline (ex : "6_CHIR - " ou "4_ANES.REA - ")
        # Le séparateur entre code et libellé est " - " (avec espaces)
        if " - " in terrain_brut:
            intitule = terrain_brut.split(" - ", 1)[1].strip()
        else:
            intitule = terrain_brut.strip()

        classement[rang] = intitule

    return classement


# ── Remplissage du fichier Excel ──────────────────────────────────────────────

def remplir_excel(classement: dict[int, str], chemin_modele: Path) -> Path:
    """
    Copie le modèle Excel et remplit la colonne 'Rang' pour chaque terrain
    trouvé dans le classement.  Le fichier résultat est enregistré dans le
    dossier de téléchargements de l'utilisateur (ou le bureau en fallback).

    Retourne le chemin du fichier généré.
    """
    # Destination : dossier Téléchargements
    dossier_dl = Path.home() / "Downloads"
    if not dossier_dl.exists():
        dossier_dl = Path.home() / "Desktop"
    if not dossier_dl.exists():
        dossier_dl = Path.home()

    chemin_sortie = dossier_dl / "voeux_campagne_rempli.xlsx"

    shutil.copy2(chemin_modele, chemin_sortie)

    wb = openpyxl.load_workbook(chemin_sortie)
    ws = wb.active  # feuille "Modèle de Vœux"

    # Ligne 1 = en-tête ; colonne A = Rang, colonne B = Intitulé du Terrain
    terrains_trouves = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        intitule_cellule = row[1].value  # colonne B
        if intitule_cellule is None:
            continue

        intitule_normalise = intitule_cellule.strip().upper()

        # Chercher cet intitulé dans le classement de l'étudiant
        for rang, intitule_voeu in classement.items():
            if intitule_voeu.upper() == intitule_normalise:
                row[0].value = rang  # colonne A = Rang
                terrains_trouves += 1
                break

    wb.save(chemin_sortie)
    return chemin_sortie, terrains_trouves


# ── Flux principal ────────────────────────────────────────────────────────────

def run():
    print("=== Récupérateur de Vœux StageTrek ===\n")
    ensure_browser_installed()

    chemin_modele = choisir_modele_excel()
    print(f"Modèle Excel sélectionné : {chemin_modele}\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=50)
        context = browser.new_context()
        page = context.new_page()

        # Étape 1 – ouvrir la page de liste des stages
        page.goto("https://stagetrek.univ-tours.fr/mes-stages")

        pause(
            "\n>>> Connecte-toi dans la fenêtre du navigateur.\n"
            ">>> Navigue ensuite jusqu'à ta page de stage (ex : /mon-stage/1212)\n"
            "    et clique sur l'onglet 'Préférences' pour t'assurer\n"
            "    que le tableau des préférences est visible.\n"
            ">>> Reviens ici et appuie sur Entrée pour LANCER LA RÉCUPÉRATION.\n"
        )

        # Étape 2 – récupérer le HTML de la page courante et extraire le classement
        print("Récupération du classement en cours...")
        html = page.content()
        classement = extraire_classement(html)

        if not classement:
            print(
                "\n⚠  Aucun vœu trouvé dans le tableau de préférences.\n"
                "   Vérifie que tu es bien sur la page de ton stage et que\n"
                "   l'onglet 'Préférences' est sélectionné."
            )
            browser.close()
            pause_avant_fermeture()
            return

        print(f"\n✔  {len(classement)} vœu(x) récupéré(s) :")
        for rang in sorted(classement):
            print(f"   Rang {rang:3d} → {classement[rang]}")

        browser.close()

    # Étape 3 – remplir le fichier Excel
    print("\nRemplissage du fichier Excel...")
    chemin_sortie, terrains_trouves = remplir_excel(classement, chemin_modele)

    # Étape 4 – informer l'étudiant
    non_trouves = len(classement) - terrains_trouves
    print(f"\n✔  Fichier généré : {chemin_sortie}")
    print(f"   {terrains_trouves} terrain(s) mis à jour dans le fichier Excel.")
    if non_trouves > 0:
        print(
            f"   ⚠  {non_trouves} intitulé(s) de vœu n'ont pas été trouvés dans\n"
            f"      le modèle Excel (différence d'orthographe possible)."
        )

    # Ouvrir le fichier automatiquement sur Mac et Windows
    try:
        if sys.platform == "darwin":
            os.system(f'open "{chemin_sortie}"')
        elif sys.platform == "win32":
            os.startfile(str(chemin_sortie))
    except Exception:
        pass  # non bloquant

    pause_avant_fermeture(
        "\nLe fichier Excel a été enregistré dans ton dossier Téléchargements.\n"
        "Appuie sur Entrée pour fermer ce programme."
    )


def main():
    try:
        run()
    except SystemExit:
        raise
    except Exception:
        print("\nUne erreur inattendue est survenue :\n")
        traceback.print_exc()
        pause_avant_fermeture("\nAppuie sur Entrée pour fermer.")
        sys.exit(1)


if __name__ == "__main__":
    main()

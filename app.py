"""
app.py
--------------
Automatise la saisie des voeux de stage sur stagetrek.univ-tours.fr
à partir d'un fichier .xlsx de récépissé (colonnes "Rang" / "Intitulé du Terrain").
"""

import sys
import os
import traceback
import re
from pathlib import Path

# IMPORTANT : doit être fait AVANT d'importer playwright.
# Force le navigateur dans un dossier permanent (sinon il faudrait le
# retélécharger à chaque lancement depuis un exécutable PyInstaller).
_DOSSIER_NAVIGATEURS = str(Path.home() / ".cache" / "auto-voeux-stagetrek" / "playwright-browsers")
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = _DOSSIER_NAVIGATEURS

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

URL_PAGE = "https://stagetrek.univ-tours.fr/preferences/modifier/1212"

def pause_avant_fermeture(message="\nAppuie sur Entrée pour fermer cette fenêtre..."):
    try:
        input(message)
    except EOFError:
        pass


def ensure_browser_installed():
    print("Vérification du navigateur nécessaire (Chromium)...")
    print("(Peut prendre plusieurs minutes la première fois : téléchargement d'environ 150 Mo.)")
    try:
        from playwright.__main__ import main as playwright_cli_main
    except Exception as e:
        raise RuntimeError(f"Impossible de préparer Playwright : {e}")

    old_argv = sys.argv
    exit_code = 0
    try:
        sys.argv = ["playwright", "install", "chromium"]
        try:
            playwright_cli_main()
        except SystemExit as e:
            exit_code = e.code
    finally:
        sys.argv = old_argv

    if exit_code not in (0, None):
        raise RuntimeError(
            f"L'installation du navigateur a échoué (code {exit_code}). "
            "Vérifie ta connexion internet puis relance le programme."
        )
    print("Navigateur prêt.")


def choisir_fichier_xlsx():
    if len(sys.argv) >= 2:
        return sys.argv[1]

    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    messagebox.showinfo(
        "Auto Vœux Stagetrek",
        "Choisis ton fichier récépissé (.xlsx) dans la fenêtre suivante.",
    )
    chemin = filedialog.askopenfilename(
        title="Choisis ton fichier récépissé (.xlsx)",
        filetypes=[("Fichiers Excel", "*.xlsx")],
    )
    root.destroy()

    if not chemin:
        print("Aucun fichier sélectionné. Fin du programme.")
        pause_avant_fermeture()
        sys.exit(0)

    return chemin


def load_wishes(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "Rang":
            header_row = i
            break
    if header_row is None:
        raise ValueError("Impossible de trouver la ligne d'en-tête 'Rang' dans le fichier.")

    wishes = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[0] is None:
            break
        rang, intitule = row[0], row[1]
        wishes.append((int(rang), str(intitule).strip()))

    wishes.sort(key=lambda x: x[0])
    return wishes


def add_one_wish(page, rang, intitule):
    # Les timeout ont été passés de 20s à 5 minutes (300000 ms) pour pallier aux lenteurs du site
    TIMEOUT_LENT = 300000 

    page.click("a.ajax-modal[data-event='event-ajouter-preference']")

    page.wait_for_selector("#rang", state="attached", timeout=TIMEOUT_LENT)
    page.wait_for_selector("#terrainStage", state="attached", timeout=TIMEOUT_LENT)

    page.fill("#rang", str(rang))
    
    # 1. Clique sur le bouton du menu déroulant (utilisation de data-id qui est fixe)
    page.click("button[data-id='terrainStage']")
    
    # 2. Cible le champ de recherche lié spécifiquement à ce select
    champ_recherche = page.locator("//div[contains(@class, 'bootstrap-select') and .//select[@id='terrainStage']]//input[@type='search']")
    champ_recherche.wait_for(state="visible", timeout=TIMEOUT_LENT)
    
    # 3. Écrit l'intitulé du voeu
    champ_recherche.fill(intitule)
    
    # 4. Laisse un temps généreux à l'interface pour filtrer la liste si l'ordinateur/navigateur rame
    page.wait_for_timeout(1000)
    
    # 5. Valide avec la touche Entrée
    champ_recherche.press("Enter")

    try:
        # Le rafraîchissement suite à la soumission peut être extrêmement lent
        with page.expect_navigation(timeout=TIMEOUT_LENT):
            page.click("#submit")
    except PWTimeout:
        try:
            pop_up_encore_ouverte = page.locator(".modal.show").count() > 0
        except Exception:
            pop_up_encore_ouverte = False
        if pop_up_encore_ouverte:
            raise RuntimeError(
                "Le site est anormalement lent ou une erreur a empêché la validation. "
                "La pop-up est restée ouverte après 3 minutes d'attente."
            )
            
    # On attend que le bouton "Ajouter une préférence" de la page principale soit de nouveau "attaché"
    # Cela garantit que la page a bien fini de recharger, sans bloquer sur des requêtes parasites en arrière-plan
    try:
        page.wait_for_selector("a.ajax-modal[data-event='event-ajouter-preference']", state="attached", timeout=TIMEOUT_LENT)
    except PWTimeout:
        pass


def run():
    print("=== Auto Vœux Stagetrek ===\n")

    ensure_browser_installed()

    chemin_fichier = choisir_fichier_xlsx()
    wishes = load_wishes(chemin_fichier)
    print(f"\n{len(wishes)} vœux détectés dans le fichier Excel.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=300000)
        context = browser.new_context()
        page = context.new_page()
        page.goto(URL_PAGE)

        pause_avant_fermeture(
            "\n>>> Connecte-toi dans la fenêtre du navigateur qui vient de s'ouvrir.\n"
            ">>> Une fois arrivé(e) sur la page 'Mes préférences' (celle avec le bouton 'Ajouter'),\n"
            ">>> reviens ici et appuie sur Entrée pour lancer la saisie automatique.\n"
        )
        
        # --- DÉTECTION DE REPRISE / VŒUX DÉJÀ SAISIS ---
        nb_deja_saisis = 0
        try:
            compteur_loc = page.locator("//div[strong[contains(text(), 'Nombre de préférence')]]")
            compteur_loc.wait_for(state="attached", timeout=5000)
            texte_compteur = compteur_loc.inner_text()
            
            # Recherche "X / 300" avec une regex
            match = re.search(r"(\d+)\s*/", texte_compteur)
            if match:
                nb_deja_saisis = int(match.group(1))
                if nb_deja_saisis > 0:
                    print(f"\n[Info Reprise] Le script a détecté que {nb_deja_saisis} vœux sont déjà enregistrés sur Stagetrek.")
                    print("               Les vœux correspondants seront ignorés pour reprendre à la suite.")
        except Exception:
            # S'il ne trouve pas le texte, il assume qu'aucun voeu n'est saisi
            pass
        # ----------------------------------------------

        succes, echecs = 0, []
        for rang, intitule in wishes:
            
            if rang <= nb_deja_saisis:
                print(f"-> Vœu de rang {rang} ignoré (déjà présent sur le site).")
                continue
                
            print(f"-> Ajout du vœu de rang {rang} : {intitule}")
            try:
                add_one_wish(page, rang, intitule)
                succes += 1
            except Exception as e:
                print(f"   ECHEC pour '{intitule}' ({e})")
                echecs.append((rang, intitule))
                try:
                    if page.locator(".modal.show").count() > 0:
                        page.click(".modal.show button.btn-secondary")
                except Exception:
                    pass
                pause_avant_fermeture(
                    "   -> L'attente maximale (3 minutes) a été dépassée ou une erreur imprévue est survenue.\n"
                    "      Corrige manuellement ce vœu dans le navigateur si besoin,\n"
                    "      ferme la pop-up si elle est encore ouverte, puis appuie sur Entrée pour continuer..."
                )

        print(f"\nTerminé : {succes} vœux ajoutés, {len(echecs)} à vérifier/corriger.")
        if echecs:
            print("Vœux en échec :")
            for rang, intitule in echecs:
                print(f"  - rang {rang} : {intitule}")

        pause_avant_fermeture(
            "\nAppuie sur Entrée pour fermer le navigateur (vérifie d'abord ton classement final sur le site)."
        )
        browser.close()


def main():
    try:
        run()
    except SystemExit:
        raise
    except Exception:
        print("\nUne erreur inattendue est survenue :\n")
        traceback.print_exc()
        pause_avant_fermeture("\nAppuie sur Entrée pour fermer.")
        sys.exit(1)


if __name__ == "__main__":
    main()
